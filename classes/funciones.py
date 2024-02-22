import ifcopenshell
import requests
import math
import os
import openpyxl
import pandas as pd
from passwords import *


#--------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------
# Classes
# -------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------


class Airtable:
    def __init__(self, token, base_id):
        self.token = token
        self.base_id = base_id

    def list(self, table, max_records=None, view='API',fields=None,filter=None,sort=None):
        url = f'https://api.airtable.com/v0/{self.base_id}/{table}'
        params = {
            'maxRecords': max_records,
            'view': view
        }
        if filter is not None:
            params['filterByFormula'] = filter
        if fields is not None:
            params['fields[]'] = fields
        if sort is not None:
            params['sort[]'] = sort
        data = []

        while True:
            response = requests.get(url, params=params, headers={'Authorization': 'Bearer ' + self.token})
            #print (response.url)
            response_data = response.json()
            data.extend(response_data.get('records', []))
            offset = response_data.get('offset', None)
            #print (offset)
            if not offset:
                break
            params['offset'] = offset
        
        depurated_data=[]
        
        for i in data:
            depurated_data.append(i['fields'])
            
        return depurated_data
        # example with custom filter: at.list('Contracts',filter="AND({Contract Type}='Obra' , {Project}='5')")

    def insert(self, table, data):
        url = f'https://api.airtable.com/v0/{self.base_id}/{table}'
        response = requests.post(url, json={'fields': data}, headers={'Authorization': 'Bearer ' + self.token})
        return response.json()

class Box:
    def __init__(self, parentjoint_id, box_type, joints3_type, joint_type, connectiongroup_type, length, corematgroupslist, q1matgroupslist, q2matgroupslist, q3matgroupslist, q4matgroupslist, floor, revitguid,c01id,c02id):
        self.id=parentjoint_id
        self.box_type=box_type
        self.joints3_type=joints3_type
        self.joint_type=joint_type
        self.corematgroups=corematgroupslist
        self.q1matgroups=q1matgroupslist
        self.q2matgroups=q2matgroupslist
        self.q3matgroups=q3matgroupslist
        self.q4matgroups=q4matgroupslist
        self.connectiongroup_type=connectiongroup_type
        self.length=length
        self.floor=floor
        self.nrbalconies=0
        self.revitguid=revitguid
        self.c01id=c01id
        self.c02id=c02id
        self.nonstructuralcost=0
        self.inferredconnectioncost=0
        self.modeledconnectioncost=0
        
class Connectiongroup:
    def __init__(self, id, description, cgclass, rl_cgtype_ctype=None, screwlong=None, screwcadence=None, anglecadence=None, angletype=None, endHD=None, balconyHD=None):
        self.id=id
        self.description=description
        self.cgclass=cgclass
        self.screwlong=screwlong
        self.screwcadence=screwcadence
        self.anglecadence=anglecadence
        self.angletype=angletype
        self.endHD=endHD
        self.balconyHD=balconyHD
        self.connectiontypes=[]
        self.rl_cgtype_ctype=rl_cgtype_ctype
        
class Connectiontype:
    def __init__(self, id, description, boxtype, is_modeled, rl_cgtype_ctype=None):
        self.id=id
        self.description=description
        self.boxtype=boxtype
        self.is_modeled=is_modeled
        self.rl_cgtype_ctype=rl_cgtype_ctype
        self.connectionlayers=[]
        self.cost=0
        
class Connectionlayer:
    def __init__(self, id, connectiontype, material_sku, performance, calcformula, fase):
        self.id=id
        self.connectiontype=connectiontype
        self.material_sku=material_sku
        self.performance=performance
        self.calcformula=calcformula
        self.fase=fase
        self.material=None
        self.cost=0
       
class Material:
    def __init__(self, sku, description, unit, cost, type):
        self.sku=sku
        self.description=description
        self.unit=unit
        self.cost=cost
        self.type=type

class Matgroup:
    def __init__(self,id,description):
        self.id=id
        self.description=description
        self.matlayers=[]

class Matlayer:
    def __init__(self,id,matgroupid,material_sku,performance,calcformula,fase):
        self.id=id
        self.matgroupid=matgroupid
        self.material_sku=material_sku
        self.performance=performance
        self.calcformula=calcformula
        self.fase=fase
        self.material=None
        self.cost=0

class Ifc:
    def __init__(self,ruta):
        self.originalruta=ruta
        self.ruta=self.originalruta.replace('\\','/')
        self.openedifc=self.open()
    
    def open(self):
        print('Abriendo IFC...')
        openedifc=ifcopenshell.open(self.ruta)
        return openedifc
    
    def list_bytype(self, element_type):
        # print('Buscando elementos en el IFC por tipo...')
        elementsbytype=self.openedifc.by_type(element_type)
        psets_elements=[]
        for i in elementsbytype:
            psets_element=ifcopenshell.util.element.get_psets(i)
            psets_elements.append(psets_element)
        return psets_elements
    
    def list_byvalue(self,element_type,pset,parameter,value):
        # print('Buscando elementos concretos en el IFC...')
        elements_byvalue=[]
        psets_elements=self.list_bytype(element_type)
        for i in psets_elements:
            if pset in i:
                parameter_value=i[pset].get(parameter,'')
                if value in parameter_value:
                    elements_byvalue.append(i)
        return elements_byvalue

class Modeledconnection:
    def __init__(self,connectiontype,revitguid,parent,floor):
        self.connectiontype=connectiontype
        self.revitguid=revitguid
        self.parent=parent
        self.floor=floor
        self.performance=1
        self.cgtype=''

        
class Bomconnectionline:
    def __init__(self,parent,performance,calcformula,fase,description,connectiontype,sku,materialcost,unit,ismodeled,quantity,layercost,cgtype,materialtype,floor,revitguid):
        self.parent=parent
        self.performance=performance
        self.calcformula=calcformula
        self.fase=fase
        self.description=description
        self.ctype=connectiontype
        self.sku=sku
        self.materialcost=materialcost
        self.unit=unit
        self.ismodeled=ismodeled
        self.quantity=quantity
        self.layercost=layercost
        self.cgtype=cgtype
        self.floor=floor
        self.materialtype=materialtype
        self.revitguid=revitguid
        
class Bomjointline:
    def __init__(self,parent,performance,calcformula,fase,description,jointtype,sku,materialcost,unit,quantity,layercost,floor,materialtype):
        self.parent=parent
        self.performance=performance
        self.calcformula=calcformula
        self.fase=fase
        self.description=description
        self.jointtype=jointtype
        self.sku=sku
        self.materialcost=materialcost
        self.unit=unit
        self.quantity=quantity
        self.layercost=layercost
        self.floor=floor
        self.materialtype=materialtype
      
      
      
      
       
#--------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------
#                                         AIRTABLE
#--------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------
     
def airtable_conection(api_key,base_id, base_name):
    print(f'  --  Ejecutando conexión a base de Airtable {base_name}')  
    connection=Airtable(api_key,base_id)        
    return connection     

JOINTS3PLAYGROUND_CONNECT=airtable_conection(adri_joints3playground_api_key,joints3_base_id, 'Joints 3 Playground')   
JOINTSPLAYGROUND_CONNECT=airtable_conection(adri_jointsplayground_api_key,joints2_base_id, 'Joints Playground')

def getrecords(connection,table,fields,view='API'):
    records=connection.list(table,fields,view)
    return records

#--------------------------------------------------------------------------------------------
# Instanciado de objetos airtable
#--------------------------------------------------------------------------------------------

def instanciar_cgtype(connectiongroups):
    print('  --  Instanciando ConnectionGroup_types...')
    connectiongroup_objects=[]
    for connectiongroup_dict in connectiongroups:
        connectiongroup_object=Connectiongroup(id=connectiongroup_dict.get('cgtype_id'),
                                        description=connectiongroup_dict.get('Description'),
                                        cgclass=connectiongroup_dict.get('api_ConnectionGroup_Class'),
                                        rl_cgtype_ctype=connectiongroup_dict.get('RL_ConnectionGroupType_ConnectionType_api',None),
                                        screwlong=connectiongroup_dict.get('param_screwlong',None),
                                        screwcadence=connectiongroup_dict.get('param_screwcadence',None),
                                        anglecadence=connectiongroup_dict.get('param_anglecadence',None),
                                        angletype=connectiongroup_dict.get('param_angletype',None),
                                        endHD=connectiongroup_dict.get('param_endHD',None),
                                        balconyHD=connectiongroup_dict.get('param_balconyHD',None))
        connectiongroup_objects.append(connectiongroup_object)
    return connectiongroup_objects

def instanciar_ctype(connectiontypes):
    print('  --  Instanciando Connection_types...')
    connectiontype_objects=[]
    for connectiontype_dict in connectiontypes:
        connectiontype_object=Connectiontype(connectiontype_dict['connection_type_id'],
                                            connectiontype_dict['description'],
                                            connectiontype_dict.get('box_type',''),
                                            connectiontype_dict['is_modeled'],
                                            connectiontype_dict.get('RL_ConnectionGroupType_Connectiontype_api',None)
                                            )
        connectiontype_objects.append(connectiontype_object)
    return connectiontype_objects

def instanciar_clayer(connectionlayers):
    print('  --  Instanciando Connection_layers...')
    connectionlayers_objects=[]
    for i in connectionlayers:
        connectionlayer_object=Connectionlayer(i.get('Name',''),
                                               i.get('connection_type_code',''),
                                               i.get('material_id',''),
                                               i.get('Performance',''),
                                               i.get('Calculation Formula',''),
                                               i.get('Fase',''))
        connectionlayers_objects.append(connectionlayer_object) 
    return connectionlayers_objects                                         
                            
def instanciar_materials(materials):
    print('  --  Instanciando materiales...')
    materials_objects=[]
    for i in materials:
        material_object=Material(i['SKU'],
                                 i['Description'],
                                 i['Measurement_Unit'],
                                 i['Estimated Cost'],
                                 i['Type of material'])
        materials_objects.append(material_object)
    return materials_objects

def instanciar_matgroups(matgroups):
    print('  --  Instanciando MatGroups...')
    matgroups_objects=[]
    for i in matgroups:
        matgroup_object=Matgroup(i['material_group'],
                                 i['Description'])
        matgroups_objects.append(matgroup_object)
    return matgroups_objects

def instanciar_matlayers(matlayers):
    print('  --  Instanciando MatLayers...')
    matlayers_objects=[]
    for i in matlayers:
        matlayer_object=Matlayer(i['material_layer'],
                                 i['material_group_id'],
                                 i['material_id'],
                                 i['Performance'],
                                 i['Calculation Formula'],
                                 i.get('Fase',''))
        matlayers_objects.append(matlayer_object)
    return matlayers_objects

def instanciar_joints(joints):
    print('  --  Instanciando Joints (Joints 2 version)...')
    joints_objects=[]
    for i in joints:
        joint_object=Matgroup(i['joint_type_id'],
                                 i.get('joint_description [toDeprecate]',''))
        joints_objects.append(joint_object)
    return joints_objects

def instanciar_jointslayers(jointslayers):
    print('  --  Instanciando Layers de Joints (Joints 2 version)...')
    jointlayers_objects=[]
    for i in jointslayers:
        jointlayer_object=Matlayer(i['Name'],
                                 i['api_id'],
                                 i['material_id'],
                                 i['Performance'],
                                 i['Calculation Formula'],
                                 i.get('Fase',''))
        jointlayers_objects.append(jointlayer_object)
    return jointlayers_objects


#--------------------------------------------------------------------------------------------
# funciones para completar objetos con layers
#--------------------------------------------------------------------------------------------

def ctypetocgtype(connectiongroup_objects,relations,connectiontype_objects):
    print('  --  Asignando los ConnectionTypes a los Connectiongroups...')
    for cg in connectiongroup_objects:
        ctypesincg=[]
        for relation in relations:
            if relation['connectiongroup_type_id']==cg.id:
                ctypedata={}
                for i in connectiontype_objects:
                    if i.id==relation['connection_type']:                        
                        ctypedata['connection_type']=i
                        ctypedata['connection_type_id']=i.id
                ctypedata['Performance']=relation['Performance']
                ctypedata['Calculation Formula']=relation['Calculation Formula']
                ctypesincg.append(ctypedata)
        cg.connectiontypes.extend(ctypesincg)
        
def clayertoctype(connectiontype_objects,connectionlayers_objects):
    print('  --  Asignando ConnectionLayers a los ConnectionTypes...')
    for ct in connectiontype_objects:
        clayersinct=[]
        for layer in connectionlayers_objects:
            if layer.connectiontype == ct.id:
                clayersinct.append(layer)
        ct.connectionlayers.extend(clayersinct)
        
def materialtoclayer(connectionlayers_objects,material_objects):
    print('  --  Asignando materiales a las ConnectionLayers...')
    for layer in connectionlayers_objects:
        for material in material_objects:
            if material.sku==layer.material_sku:
                layer.material=material

def matlayertomatgroup(matgroup_objects,matlayer_objects):
    print('  --  Asignando MatLayers a Matgroups...')
    for matgroup in matgroup_objects:
        matlayersinmatgroup=[]
        for matlayer in matlayer_objects:
            if matlayer.matgroupid==matgroup.id:
                matlayersinmatgroup.append(matlayer)
        matgroup.matlayers.extend(matlayersinmatgroup)
        
def materialtomatlayer(matlayer_objects,material_objects):
    print('  --  Asignando materiales a las MatLayers...')
    for matlayer in matlayer_objects:
        for material in material_objects:
            if material.sku==matlayer.material_sku:
                matlayer.material=material

def jointlayertojoint(joints_objects, jointlayers_objects):
    print('  --  Asignando Layers a Joints...')
    for joint in joints_objects:
        layersinjoint=[]
        for layer in jointlayers_objects:
            if layer.matgroupid==joint.id:
                layersinjoint.append(layer)
        joint.matlayers.extend(layersinjoint)
        
def materialtojointlayer(jointlayer_objects,material_objects):
    print('  --  Asignando materiales a Joint Layers...')
    for layer in jointlayer_objects:
        for material in material_objects:
            if material.sku==layer.material_sku:
                layer.material=material

#--------------------------------------------------------------------------------------------
# funciones para completar objetos de Airtable con costes
#--------------------------------------------------------------------------------------------

def layercost(connectionlayer_objects):
    print('  --  Calculando coste de las ConnectionLayers...')
    for i in connectionlayer_objects:
        if i.material is not None:
            i.cost=i.performance*i.material.cost
            
def connectioncost(connectiontype_objects):
    print('  --  Calculando coste de las ConnectionTypes...')
    for connection in connectiontype_objects:        
        for layer in connection.connectionlayers:            
            connection.cost=connection.cost+layer.cost

















            
#--------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------
#                                           IFC
#--------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------


def get_allboxes(ifc_object):
    type='IfcBuildingElementProxy'
    pset='EI_Elements Identification'
    parameter='EI_Type'
    value='Joint'       
    boxes=ifc_object.list_byvalue(type,pset,parameter,value)
    boxes_info=[]
    for i_psets in boxes:        
        JS_Joint_Specification = i_psets.get('JS_Joint Specification', {})
        QU_Quantity=i_psets.get('QU_Quantity',{})
        Pset_QuantityTakeOff=i_psets.get('Pset_QuantityTakeOff',{})
        EI_Elements_Identification=i_psets.get('EI_Elements Identification',{})
        EI_LocalisationCodeFloor=EI_Elements_Identification.get('EI_LocalisationCodeFloor','')
        JointTypeID = JS_Joint_Specification.get('JS_JointTypeID', '')
        cgt=JS_Joint_Specification.get('JS_ConnectionGroupTypeID', '')
        cgc=JS_Joint_Specification.get('JS_ConnectionGroupClass','')
        parent_joint_id = JS_Joint_Specification.get('JS_ParentJointInstanceID', '')
        r_guid = i_psets['EI_Interoperability'].get('RevitGUID', '')
        QU_Length=round(QU_Quantity.get('QU_Length_m', 0), 2)
        Box_type=Pset_QuantityTakeOff.get('Reference', '')
        joint_type=JS_Joint_Specification.get('JS_JointType','')
        inst_a=JS_Joint_Specification.get('JS_C01_ID', '')
        inst_b=JS_Joint_Specification.get('JS_C02_ID', '')
        corematgroup=JS_Joint_Specification.get('JS_CORE_matgroup','')
        Q1matgroup=JS_Joint_Specification.get('JS_Q1_matgroup','')
        Q2matgroup=JS_Joint_Specification.get('JS_Q2_matgroup','')
        Q3matgroup=JS_Joint_Specification.get('JS_Q3_matgroup','')
        Q4matgroup=JS_Joint_Specification.get('JS_Q4_matgroup','')
        parameters_info={'RevitGUID':r_guid,
                         'JS_ParentJointInstanceID':parent_joint_id,
                         'JS_JointTypeID':JointTypeID,
                         'JS_ConnectionGroupTypeID':cgt,
                         'Core Matgroup':corematgroup,
                         'Q1 Matgroup':Q1matgroup,
                         'Q2 Matgroup':Q2matgroup,
                         'Q3 Matgroup':Q3matgroup,
                         'Q4 Matgroup':Q4matgroup,
                         'QU_Length_m':QU_Length,
                         'Box_type':Box_type,
                         'JS_JointType':joint_type,
                         'JS_ConnectionGroupClass':cgc,
                         'JS_C01_ID':inst_a,
                         'JS_C02_ID':inst_b,
                         'EI_LocalisationCodeFloor':EI_LocalisationCodeFloor
                         }
        boxes_info.append(parameters_info)
    return boxes_info


def filter_boxes(all_boxes):
    parents_unicos=set()
    filtered_boxes=[]
    print('  --  Filtrando cajas...')
    
    for info in all_boxes:
        parent=info['JS_ParentJointInstanceID']
        if parent not in parents_unicos and info['JS_JointType']!='':
            parents_unicos.add(parent)
            filtered_boxes.append(info)
    
    # print('Primer barrido listo')
            
    for info in all_boxes:
        parent=info['JS_ParentJointInstanceID']
        if parent not in parents_unicos and 'GMO' in info['Box_type']:
            parents_unicos.add(parent)
            filtered_boxes.append(info)
    
    # print('Segundo barrido listo')
    
    for info in all_boxes:
        parent=info['JS_ParentJointInstanceID']
        if parent not in parents_unicos:
            parents_unicos.add(parent)
            filtered_boxes.append(info) 
            
    # print('Tercer y último barrido listo')
                
    return filtered_boxes


def get_balconys(ifc_object):
    balconys=[]
    elements_bytype=ifc_object.list_bytype('IfcWindow')
    for i_psets in elements_bytype:
        QU_Quantity = i_psets.get('QU_Quantity', {})
        EI_Elements_Identification = i_psets.get('EI_Elements Identification', {})
        EI_OpeningType = EI_Elements_Identification.get('EI_OpeningType', '')
        EI_HostComponentInstanceID=EI_Elements_Identification.get('EI_HostComponentInstanceID', '')
        QU_Height=QU_Quantity.get('QU_Height_m', '')
        if QU_Height>1.8:
            balconys.append({'EI_OpeningType':EI_OpeningType,'EI_HostComponentInstanceID':EI_HostComponentInstanceID,'QU_Height_m':QU_Height})
    return balconys

def get_huecosdepaso(ifc_object):
    huecos_paso=[]
    elements_bytype=ifc_object.list_bytype('IfcWall')
    for i_psets in elements_bytype:
        Pset_ProductRequirements=i_psets.get('Pset_ProductRequirements',{}) 
        EI_Elements_Identification=i_psets.get('EI_Elements Identification',{})   
        EI_HostComponentInstanceID=EI_Elements_Identification.get('EI_HostComponentInstanceID','')
        EI_OpeningType=EI_Elements_Identification.get('EI_OpeningType', '') 
        Category=Pset_ProductRequirements.get('Category', '')
        Params_info={'EI_HostComponentInstanceID':EI_HostComponentInstanceID,'EI_OpeningType':EI_OpeningType, 'Category':Category}
        if Category=='Doors':
                huecos_paso.append(Params_info)
    return huecos_paso

def get_doors(ifc_object):
    doors=[]
    elements_bytype=ifc_object.list_bytype('IfcDoor')
    for i_psets in elements_bytype:
        Pset_ProductRequirements=i_psets.get('Pset_ProductRequirements',{}) 
        EI_Elements_Identification=i_psets.get('EI_Elements Identification',{})
        EI_HostComponentInstanceID=EI_Elements_Identification.get('EI_HostComponentInstanceID','')
        EI_OpeningType=EI_Elements_Identification.get('EI_OpeningType', '') 
        Category=Pset_ProductRequirements.get('Category', '')
        Params_info={'EI_HostComponentInstanceID':EI_HostComponentInstanceID,
                     'EI_OpeningType':EI_OpeningType,
                     'Category':Category}
        doors.append(Params_info)
    return doors

def get_modeledconnections(ifc_object):
    modeled_connections=[]
    elements_byvalue=ifc_object.list_byvalue('IfcBuildingElementProxy','EI_Elements Identification','EI_Type','Connection')
    for i_psets in elements_byvalue:
        JS_Joint_Specification = i_psets.get('JS_Joint Specification', {})    
        EI_Interoperability=i_psets.get('EI_Interoperability', {})
        EI_Elements_Identification=i_psets.get('EI_Elements Identification',{})
        EI_LocalisationCodeFloor=EI_Elements_Identification.get('EI_LocalisationCodeFloor','')
        connectiontype_id = JS_Joint_Specification.get('JS_ConnectionTypeID', '')        
        parent_joint_id = JS_Joint_Specification.get('JS_ParentJointInstanceID', '')
        r_guid = EI_Interoperability.get('RevitGUID', '')      
        parameters_info={'RevitGUID':r_guid,
                         'JS_ParentJointInstanceID': parent_joint_id,
                         'JS_ConnectionTypeID':connectiontype_id,
                         'EI_LocalisationCodeFloor':EI_LocalisationCodeFloor
                         }
        modeled_connections.append(parameters_info)
    return modeled_connections
        

#agrupamos get_doors y get_huecosdepaso
def group_alldoors(ifc_object):
    alldoors_info=get_doors(ifc_object)+get_huecosdepaso(ifc_object)
    return alldoors_info 

#función que llama a recopilar todas los huecos que nos interesan (junta las 3 funciones huecosdepaso, doors y balconies)
def counter_openings_byinstance(ifc_object):
    print('  --  Extrayendo y calculando huecos del modelo que requieren HoldDowns...')
    #buscamos una respusesta tipo: [{ host_comp_instance : nr_huecosacontabilizar } , { host_comp_instance : nr_huecosacontabilizar }...]
    
    info_ifcwindow=get_balconys(ifc_object)
    alldoors_info=group_alldoors(ifc_object)
    recuento_balconerasporinstancia={}
    
    for balcony in info_ifcwindow:
        instancia=balcony['EI_HostComponentInstanceID']
        if instancia in recuento_balconerasporinstancia:
            recuento_balconerasporinstancia[instancia] += 1
        else:
            recuento_balconerasporinstancia[instancia] = 1
            
    for door in alldoors_info:
        instancia=door['EI_HostComponentInstanceID']
        if instancia in recuento_balconerasporinstancia:
            recuento_balconerasporinstancia[instancia] += 1
        else:
            recuento_balconerasporinstancia[instancia] = 1
    
    return recuento_balconerasporinstancia



#--------------------------------------------------------------------------------------------
# funciones de instanciado de los objetos procedentes del archivo IFC
#--------------------------------------------------------------------------------------------

def instanciarIFC(ruta):
    # print('  --  Instanciando IFC...')
    ifc_object=Ifc(ruta) 
    return ifc_object

def instanciarboxes(boxes,materialgroup_objects,joints2_objects,connectiongroups_objects):
    print('  --  Instanciando cajas...')
    boxes_objects=[]
    for i in boxes:
        parentjoint_id=i.get('JS_ParentJointInstanceID','')
        box_type=i.get('Box_type','')
        joints3_type=i.get('JS_JointType','')
        
        joint_type=''
        for joint in joints2_objects:
            joint_type_id=i.get('JS_JointTypeID','')
            if joint.id==joint_type_id:
                joint_type=joint
                
        connectiongroup_type=''
        for connectiongroup in connectiongroups_objects:
            connectiongroup_type_id=i.get('JS_ConnectionGroupTypeID','')
            if connectiongroup.id==connectiongroup_type_id:
                connectiongroup_type=connectiongroup
        
        length=i['QU_Length_m']
        
        corematgroupslist=[]
        q1matgroupslist=[]
        q2matgroupslist=[]
        q3matgroupslist=[]
        q4matgroupslist=[]
        for materialgroup in materialgroup_objects:
            corematgroups_text=i.get('Core Matgroup','')
            corematgroups_id=corematgroups_text.split(',')
            if materialgroup.id in corematgroups_id:
                corematgroupslist.append(materialgroup)
                
            q1matgroups_text=i.get('Q1 Matgroup','')
            q1matgroups_id=q1matgroups_text.split(',')
            if materialgroup.id in q1matgroups_id:
                q1matgroupslist.append(materialgroup)
                
            q2matgroups_text=i.get('Q2 Matgroup','')
            q2matgroups_id=q2matgroups_text.split(',')
            if materialgroup.id in q2matgroups_id:
                q2matgroupslist.append(materialgroup)
                
            q3matgroups_text=i.get('Q3 Matgroup','')
            q3matgroups_id=q3matgroups_text.split(',')
            if materialgroup.id in q3matgroups_id:
                q3matgroupslist.append(materialgroup)
            
            q4matgroups_text=i.get('Q4 Matgroup','')
            q4matgroups_id=q4matgroups_text.split(',')
            if materialgroup.id in q4matgroups_id:
                q4matgroupslist.append(materialgroup)
                
        floor=i.get('EI_LocalisationCodeFloor','')
        revitguid=i.get('RevitGUID','')
        c01id=i.get('JS_C01_ID','')
        c02id=i.get('JS_C02_ID','')
        
        box_object=Box(parentjoint_id,
                       box_type,
                       joints3_type,
                       joint_type,
                       connectiongroup_type,
                       length,corematgroupslist,
                       q1matgroupslist,
                       q2matgroupslist,
                       q3matgroupslist,
                       q4matgroupslist,
                       floor,
                       revitguid,
                       c01id,
                       c02id)
        boxes_objects.append(box_object)
    return boxes_objects

def instanciarconexionesmodeladas(conexiones_modeladas,connectiontype_objects):
    print('  --  Instanciando herrajes modelados...')
    herrajes_objects=[]
    for i in conexiones_modeladas:
        if i['JS_ConnectionTypeID']!='':
            for connection_object in connectiontype_objects:
                if connection_object.id==i['JS_ConnectionTypeID']:
                    herraje_object=Modeledconnection(connection_object,
                                             i['RevitGUID'],
                                             i.get('JS_ParentJointInstanceID',''),
                                             i.get('EI_LocalisationCodeFloor',''))
                    herrajes_objects.append(herraje_object)
    return herrajes_objects


#--------------------------------------------------------------------------------------------
# funciones para completar objetos de IFC 
#--------------------------------------------------------------------------------------------
            
def box_nonstructural_cost(boxes_objects,bomlines_joints):
    for box in boxes_objects:
        for jointline in bomlines_joints:
            if jointline.parent==box.id:
                box.nonstructuralcost=box.nonstructuralcost+jointline.layercost        

def box_inferredconnection_cost(boxes_objects,bomlines_inferredconnections):
    for box in boxes_objects:               
        for connectionline in bomlines_inferredconnections:
            if connectionline.parent==box.id:
                box.inferredconnectioncost=box.inferredconnectioncost+connectionline.layercost
                        
def box_modeledconnection_cost(boxes_objects,bomlines_modeledconnections):
    for box in boxes_objects:
        if bomlines_modeledconnections!=[]:           
            for connectionline in bomlines_modeledconnections:
                if connectionline.parent==box.id:
                    box.modeledconnectioncost=box.modeledconnectioncost+connectionline.layercost
        else:
            box.modeledconnectioncost=''

def nropenings_to_boxes(ifc_object,boxes_objects):
    print('  --  Asignando cantidad de openings a cada caja...')
    openings_by_instance=counter_openings_byinstance(ifc_object)
    
    contador=0
    
    for box in boxes_objects:
        componente=box.c01id
        if box.box_type=="H.ST_Bottom":
            if componente!='' and componente in openings_by_instance.keys():
                box.nrbalconies=openings_by_instance[componente]
                contador=contador+openings_by_instance[componente]
                
    if contador==0:print('------------------ /!\ /!\  ALERTA: No ha sido posible encontrar ningún opening que genere holddown en el modelo /!\ /!\  ------------------')

                
#--------------------------------------------------------------------------------------------
# funciones generadoras de BOM
#--------------------------------------------------------------------------------------------

def bomlines_joints(boxes_objects):
    joints_bom_lines=[]
    for box in boxes_objects:        
        joint_type=box.joint_type
        corematgroups=box.corematgroups
        q1matgroups=box.q1matgroups
        q2matgroups=box.q2matgroups
        q3matgroups=box.q3matgroups
        q4matgroups=box.q4matgroups 
        
        jointandmatgroups=[joint_type]+corematgroups+q1matgroups+q2matgroups+q3matgroups+q4matgroups      
        
        for joint in jointandmatgroups:
            if joint != '' and joint!='J_novalue':
                layers=joint.matlayers
                for layer in layers:
                    parent=box.id
                    performance=layer.performance
                    calcformula=layer.calcformula
                    fase=layer.fase
                    description=layer.material.description
                    jointtype=joint
                    sku=layer.material.sku
                    materialcost=layer.material.cost
                    unit=layer.material.unit
                    materialtype=layer.material.type
                    
                    if calcformula=='Fix value':
                        quantity=performance
                        layercost=layer.cost
                    elif calcformula=='Length * performance':
                        quantity=performance*box.length
                        layercost=layer.cost*box.length
                                    
                    floor=box.floor
                    
                    line=Bomjointline(parent,
                                    performance,
                                    calcformula,
                                    fase,
                                    description,
                                    jointtype,
                                    sku,
                                    materialcost,
                                    unit,quantity,
                                    layercost,
                                    floor,
                                    materialtype)
                    joints_bom_lines.append(line)
    return joints_bom_lines

def bomlines_inferredconnections(boxes_objects):
    inferredconnections_bom_lines=[]
    for box in boxes_objects:
        
        connectiongrouptype=box.connectiongroup_type
        long=box.length
        nrbalconies=box.nrbalconies
        
        if connectiongrouptype != '' and connectiongrouptype!='J_novalue' and connectiongrouptype!='CG_novalue':
            connectiontypes=connectiongrouptype.connectiontypes
            if connectiontypes!=[]:
                for contype in connectiontypes:
                    ctype=contype['connection_type']
                    ctype_performance=contype['Performance']
                    ctype_calculationformula=contype['Calculation Formula']
                                    
                    if ctype_calculationformula=='Fix value':
                        ctype_multiplicador=ctype_performance
                    elif ctype_calculationformula=='Length * performance':
                        ctype_multiplicador=ctype_performance*long
                    elif ctype_calculationformula=='Opening * performance':
                        ctype_multiplicador=ctype_performance*nrbalconies
                    
                    
                    for clayer in ctype.connectionlayers:
                        parent=box.id
                        performance=clayer.performance
                        calcformula=ctype_calculationformula
                        fase=clayer.fase
                        description=clayer.material.description
                        connectiontype=ctype.id
                        sku=clayer.material.sku
                        materialcost=clayer.material.cost
                        unit=clayer.material.unit
                        ismodeled=ctype.is_modeled
                        quantity=clayer.performance*ctype_multiplicador
                        if unit=='U':
                            quantity=math.ceil(quantity)
                        layercost=clayer.material.cost*quantity 
                        cgtype=box.connectiongroup_type.id
                        materialtype=clayer.material.type
                        floor=box.floor
                        revitguid=''
                        line=Bomconnectionline(parent,
                                            performance,
                                            calcformula,
                                            fase,
                                            description,
                                            connectiontype,
                                            sku,
                                            materialcost,
                                            unit,
                                            ismodeled,
                                            quantity,
                                            layercost,
                                            cgtype,
                                            materialtype,
                                            floor,
                                            revitguid)
                        inferredconnections_bom_lines.append(line)
    return inferredconnections_bom_lines

def bomlines_modeledconnections(boxes_objects,herrajes_objects):
    modeledconnections_bom_lines=[]
    for box in boxes_objects:
        
        connectiongrouptype=box.connectiongroup_type
        long=box.length
        nrbalconies=box.nrbalconies
        
        if connectiongrouptype != '' and connectiongrouptype!='J_novalue' and connectiongrouptype!='CG_novalue' and herrajes_objects!=[]:
            connectiontypes=connectiongrouptype.connectiontypes
            if connectiontypes!=[]:
                for contype in connectiontypes:
                    ctype=contype['connection_type']
                    ctype_performance=contype['Performance']
                    ctype_calculationformula=contype['Calculation Formula']
                    ctype_ismodeled=ctype.is_modeled
                                    
                    if ctype_calculationformula=='Fix value':
                        ctype_multiplicador=ctype_performance
                    elif ctype_calculationformula=='Length * performance':
                        ctype_multiplicador=ctype_performance*long
                    elif ctype_calculationformula=='Opening * performance':
                        ctype_multiplicador=ctype_performance*nrbalconies
                    
                    
                    for clayer in ctype.connectionlayers:
                        if ctype_ismodeled=='No':
                            parent=box.id
                            performance=clayer.performance
                            calcformula=ctype_calculationformula
                            fase=clayer.fase
                            description=clayer.material.description
                            connectiontype=ctype.id
                            sku=clayer.material.sku
                            materialcost=clayer.material.cost
                            unit=clayer.material.unit
                            ismodeled=ctype.is_modeled
                            quantity=clayer.performance*ctype_multiplicador
                            if unit=='U':
                                quantity=math.ceil(quantity)
                            layercost=clayer.material.cost*quantity 
                            cgtype=box.connectiongroup_type.id
                            materialtype=clayer.material.type
                            floor=box.floor                            
                            line=Bomconnectionline(parent,
                                                performance,
                                                calcformula,
                                                fase,
                                                description,
                                                connectiontype,
                                                sku,
                                                materialcost,
                                                unit,
                                                ismodeled,
                                                quantity,
                                                layercost,
                                                cgtype,
                                                materialtype,
                                                floor,
                                                '')
                            modeledconnections_bom_lines.append(line)
    
    
    for herraje in herrajes_objects:  
        ctype=herraje.connectiontype 
        for box in boxes_objects:
            if herraje.parent==box.id and box.connectiongroup_type!='':
                herraje.cgtype=box.connectiongroup_type.id
        for clayer in ctype.connectionlayers:        
            parent=herraje.parent
            performance=clayer.performance
            calcformula='Unidad Modelada'
            fase=clayer.fase
            description=clayer.material.description
            connectiontype=ctype.id
            sku=clayer.material.sku
            materialcost=clayer.material.cost
            unit=clayer.material.unit
            ismodeled=ctype.is_modeled
            quantity=clayer.performance            
            layercost=clayer.cost
            cgtype=herraje.cgtype            
            materialtype=clayer.material.type
            floor=herraje.floor
            revitguid=herraje.revitguid
            line=Bomconnectionline(parent,
                                performance,
                                calcformula,
                                fase,
                                description,
                                connectiontype,
                                sku,
                                materialcost,
                                unit,
                                ismodeled,
                                quantity,
                                layercost,
                                cgtype,
                                materialtype,
                                floor,
                                revitguid)
            modeledconnections_bom_lines.append(line) 
    return modeledconnections_bom_lines  


















            
#--------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------
#                                           Exportación a Excel
#--------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------


                
#--------------------------------------------------------------------------------------------
# funciones para generar diccionarios que se exporten a excel
#--------------------------------------------------------------------------------------------

def boxes_to_dictlist(boxes_objects):
    lista=[]
    for box in boxes_objects:
        
        if box.joint_type!='' and box.joint_type!= 'J_novalue':
            jointtype=box.joint_type.id
        else:
            jointtype=''
            
        if box.corematgroups != []:
            corematgroups_ids=[]
            for i in box.corematgroups:
                corematgroups_ids.append(i.id)
            list_corematgroups=','.join(corematgroups_ids)
        else:
            list_corematgroups=''
        
        if box.q1matgroups != []:
            q1rematgroups_ids=[]
            for i in box.q1matgroups:
                q1rematgroups_ids.append(i.id)
            list_q1rematgroups=','.join(q1rematgroups_ids)
        else:
            list_q1rematgroups=''
        
        if box.q2matgroups != []:
            q2rematgroups_ids=[]
            for i in box.q2matgroups:
                q2rematgroups_ids.append(i.id)
            list_q2rematgroups=','.join(q2rematgroups_ids)
        else:
            list_q2rematgroups=''
            
        if box.q3matgroups != []:
            q3rematgroups_ids=[]
            for i in box.q3matgroups:
                q3rematgroups_ids.append(i.id)
            list_q3rematgroups=','.join(q3rematgroups_ids)
        else:
            list_q3rematgroups=''
            
        if box.q4matgroups != []:
            q4rematgroups_ids=[]
            for i in box.q4matgroups:
                q4rematgroups_ids.append(i.id)
            list_q4rematgroups=','.join(q4rematgroups_ids)
        else:
            list_q4rematgroups=''
        
        if box.connectiongroup_type!='' and box.connectiongroup_type!= 'J_novalue' and box.connectiongroup_type!= 'CG_novalue':
            cgtype=box.connectiongroup_type.id
        else:
            cgtype=''
        
        diccionario={'RevitGUID':box.revitguid,
                     'Parent_id':box.id,
                     'Box Type':box.box_type,
                     'Joint Type (Joints 3)':box.joints3_type,
                     'JointType_id (Joints 2)':jointtype,
                     'Core Matgroup':list_corematgroups,
                     'Q1 Matgroups':list_q1rematgroups,
                     'Q2 Matgroups':list_q2rematgroups,
                     'Q3 Matgroups':list_q3rematgroups,
                     'Q4 Matgroups':list_q4rematgroups,
                     'ConnectiongroupType_id':cgtype,
                     'Length':box.length,
                     'Nr Balconies':box.nrbalconies,
                     'Component 1':box.c01id,
                     'Component 2':box.c02id,
                     'Floor':box.floor,
                     'Non Structural Materials Cost':box.nonstructuralcost,
                     'Inferred Connections Cost':box.inferredconnectioncost,
                     'Modeled Connections Cost':box.modeledconnectioncost}
        lista.append(diccionario)
    return lista

def jointline_to_dictlist(joints_bom_lines):
    lista=[]
    for line in joints_bom_lines:
        parent=line.parent
        performance=line.performance
        calcformula=line.calcformula
        fase=line.fase
        description=line.description
        jointtype=line.jointtype.id
        sku=line.sku
        materialcost=line.materialcost
        unit=line.unit        
        quantity=line.quantity
        layercost=line.layercost        
        floor=line.floor   
        materialtype=line.materialtype            
        
        diccionario={'Parent_ID':parent,
                     'Performance':performance,
                     'JointTypeid/MaterialGroup':jointtype,
                     'Calculation Formula':calcformula,
                     'SKU':sku,
                     'Unit':unit,
                     'Description':description,
                     'Estimated Material Cost':materialcost,
                     'Fase':fase,
                     'Material Type':materialtype,
                     'Quantity':quantity,
                     'Layer Cost':layercost,
                     'Floor':floor}
        lista.append(diccionario)
    return lista

def inferredconnectionline_to_dictlist(inferredconnections_bom_lines):
    lista=[]
    for line in inferredconnections_bom_lines:
        parent=line.parent
        performance=line.performance
        calcformula=line.calcformula
        fase=line.fase
        description=line.description
        ctype=line.ctype        
        sku=line.sku
        materialcost=line.materialcost
        unit=line.unit 
        ismodeled=line.ismodeled       
        quantity=line.quantity
        layercost=line.layercost 
        cgtype=line.cgtype       
        floor=line.floor  
        materialtype=line.materialtype             
        
        diccionario={'Parent_ID':parent,
                     'Performance':performance,
                     'Connectiongroup type':cgtype,
                     'Connection Type':ctype,
                     'Is Modeled?':ismodeled,
                     'Calculation Formula':calcformula,
                     'SKU':sku,
                     'Unit':unit,
                     'Description':description,
                     'Estimated Material Cost':materialcost,
                     'Fase':fase,
                     'Material Type':materialtype,
                     'Quantity':quantity,
                     'Layer Cost':layercost,
                     'Floor':floor}
        lista.append(diccionario)
    return lista

def modeledconnectionline_to_dictlist(modeledconnections_bom_lines):
    lista=[]
    for line in modeledconnections_bom_lines:
        parent=line.parent
        performance=line.performance
        calcformula=line.calcformula
        fase=line.fase
        description=line.description
        ctype=line.ctype        
        sku=line.sku
        materialcost=line.materialcost
        unit=line.unit 
        ismodeled=line.ismodeled       
        quantity=line.quantity
        layercost=line.layercost 
        cgtype=line.cgtype       
        floor=line.floor  
        materialtype=line.materialtype
        revitguid=line.revitguid             
        
        diccionario={'RevitGUID':revitguid,
                     'Parent_ID':parent,
                     'Performance':performance,
                     'Connectiongroup type':cgtype,
                     'Connection Type':ctype,
                     'Is Modeled?':ismodeled,
                     'Calculation Formula':calcformula,
                     'SKU':sku,
                     'Unit':unit,
                     'Description':description,
                     'Estimated Material Cost':materialcost,
                     'Fase':fase,
                     'Material Type':materialtype,
                     'Quantity':quantity,
                     'Layer Cost':layercost,
                     'Floor':floor}
        lista.append(diccionario)
    return lista
                

#--------------------------------------------------------------------------------------------
# funciones para exportar archivo
#--------------------------------------------------------------------------------------------

def exportar_a_excel(boxeslist,jointslist,inferredconnectionslist,modeledconnectionslist,ruta):
    ruta_real=ruta.replace('\\','/')
    
    nombrearchivo=os.path.splitext(os.path.basename(ruta_real))[0]
    
    df = pd.DataFrame(boxeslist)
    ruta_excel = os.path.join(os.path.dirname(ruta_real), f'{nombrearchivo}.xlsx')
    df.to_excel(ruta_excel, index=False)        
    excel_file = openpyxl.load_workbook(ruta_excel)      
    excel_file.active.title="Boxes List"
    
    hoja_materiales_joints = excel_file.create_sheet(title="Non structural Materials")
    if jointslist:
        keys = list(jointslist[0].keys())
        hoja_materiales_joints.append(keys)
        for material in jointslist:
            row_data = [material[key][0] if isinstance(material.get(key, ''), list) else material.get(key, '') for key in keys]
            hoja_materiales_joints.append(row_data)
            
    hoja_inferredconnections = excel_file.create_sheet(title="Inferred Connections")
    if inferredconnectionslist:
        keys = list(inferredconnectionslist[0].keys())
        hoja_inferredconnections.append(keys)
        for herraje in inferredconnectionslist:
            row_data = [herraje[key][0] if isinstance(herraje.get(key, ''), list) else herraje.get(key, '') for key in keys]
            hoja_inferredconnections.append(row_data)
    
    if modeledconnectionslist!=[]:        
        hoja_modeledconnections = excel_file.create_sheet(title="Modeled Connections")
        if modeledconnectionslist:
            keys = list(modeledconnectionslist[0].keys())
            hoja_modeledconnections.append(keys)
            for herraje in modeledconnectionslist:
                row_data = [herraje[key][0] if isinstance(herraje.get(key, ''), list) else herraje.get(key, '') for key in keys]
                hoja_modeledconnections.append(row_data)
            
    excel_file.save(ruta_excel)
    print('')
    print('Archivo guardado con éxito')
    print('')













            
#--------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------
#                                           APP
#--------------------------------------------------------------------------------------------
#--------------------------------------------------------------------------------------------


def generate_alldata_joints_fromIFC(ruta):
    #-------------------recopilar datos de airtable
    
    print('Extrayendo datos necesarios de Airtable...')
    
    connectiongroup_types=getrecords(JOINTS3PLAYGROUND_CONNECT,cgtype_table,['cgtype_id','Description','api_ConnectionGroup_Class','param_screwlong', 'param_screwcadence','param_anglecadence','param_angletype','param_endHD','param_balconyHD', 'RL_ConnectionGroupType_ConnectionType_api'],view='AM')
    connection_types=getrecords(JOINTS3PLAYGROUND_CONNECT,connectiontype_table,['connection_type_id','description','box_type (from boxtype_id)','is_modeled','RL_ConnectionGroupType_Connectiontype_api'])
    relations=getrecords(JOINTS3PLAYGROUND_CONNECT,rl_cgtype_ctype_table,['RL_cgtype_ctype','connectiongroup_type_id','connection_type','Performance','Calculation Formula']) 
    connection_layers=getrecords(JOINTS3PLAYGROUND_CONNECT,clayers_table,['Name','connection_type_code','material_id','Performance', 'Calculation Formula','Current_material_cost','Units','Description (from Material)', 'Fase'])
    materials=getrecords(JOINTS3PLAYGROUND_CONNECT,materials_table,['SKU','Description','Measurement_Unit','Estimated Cost','Type of material'])
    materialgroups=getrecords(JOINTS3PLAYGROUND_CONNECT,materialgroups_table,['material_group','Description'])
    matlayers=getrecords(JOINTS3PLAYGROUND_CONNECT,materiallayers_table,['material_layer','material_group_id','material_id','Performance','Calculation Formula','Fase'])

    joints=getrecords(JOINTSPLAYGROUND_CONNECT,joints_table,['joint_type_id','joint_description [toDeprecate]'],view='Adri export')
    jointslayers=getrecords(JOINTSPLAYGROUND_CONNECT,jointlayers_table,['Name','api_id','material_id','Performance','Calculation Formula','Fase'],view='AMG_export')

        
    #------------------instanciado de datos de airtable
    
    print('Instanciando datos de Airtable...')
    
    connectiongroup_objects=instanciar_cgtype(connectiongroup_types)
    connectiontype_objects=instanciar_ctype(connection_types)
    connectionlayer_objects=instanciar_clayer(connection_layers)
    material_objects=instanciar_materials(materials)
    matgroup_objects=instanciar_matgroups(materialgroups)
    matlayer_objects=instanciar_matlayers(matlayers)
    joints_objects=instanciar_joints(joints)
    jointlayers_objects=instanciar_jointslayers(jointslayers)
    
    allLayers=jointlayers_objects+matlayer_objects+connectionlayer_objects
    
    #------------------asignar datos de objetos airtable
    ctypetocgtype(connectiongroup_objects,relations,connectiontype_objects)
    clayertoctype(connectiontype_objects,connectionlayer_objects)
    materialtoclayer(connectionlayer_objects,material_objects)
    matlayertomatgroup(matgroup_objects,matlayer_objects)
    materialtomatlayer(matlayer_objects,material_objects)
    jointlayertojoint(joints_objects,jointlayers_objects)
    materialtojointlayer(jointlayers_objects,material_objects)
    
    
    #------------------asignar datos de coste a layers
    layercost(allLayers)
    connectioncost(connectiontype_objects)
    
    #------------------recopilar datos del modelo
    
    ifc_object=instanciarIFC(ruta)
    
    print('  --  Leyendo IFC...')
        
    allboxes=get_allboxes(ifc_object)
    filtered_boxes=filter_boxes(allboxes)
    modeledconnections=get_modeledconnections(ifc_object)    
       
    #------------------instanciado de datos del modelo
    boxes_objects=instanciarboxes(filtered_boxes,matgroup_objects,joints_objects,connectiongroup_objects)
    herrajes_objects=instanciarconexionesmodeladas(modeledconnections,connectiontype_objects)
    
    #------------------asignar datos de ventanas a las cajas
    nropenings_to_boxes(ifc_object,boxes_objects)
    
    #------------------generar filas del BOM
    joints_bom_lines=bomlines_joints(boxes_objects)
    inferredconnections_bom_lines=bomlines_inferredconnections(boxes_objects)
    modeledconnections_bom_lines=bomlines_modeledconnections(boxes_objects,herrajes_objects)
        
    #------------------asignación de coste estimado y real a cada caja
    box_nonstructural_cost(boxes_objects,joints_bom_lines)
    box_inferredconnection_cost(boxes_objects,inferredconnections_bom_lines)
    box_modeledconnection_cost(boxes_objects,modeledconnections_bom_lines)
    
    return boxes_objects,joints_bom_lines,inferredconnections_bom_lines,modeledconnections_bom_lines
    
def transform_boxes_info_for_bom_excel_and_generate(boxes_objects,joints_bom_lines,inferredconnections_bom_lines,modeledconnections_bom_lines,ruta):
    
    print('Preparando para exportar a excel...')
    
    boxeslist=boxes_to_dictlist(boxes_objects)
    jointslist=jointline_to_dictlist(joints_bom_lines)
    inferredconnectionslist=inferredconnectionline_to_dictlist(inferredconnections_bom_lines)
    modeledconnectionslist=modeledconnectionline_to_dictlist(modeledconnections_bom_lines)
    
    exportar_a_excel(boxeslist,jointslist,inferredconnectionslist,modeledconnectionslist,ruta)